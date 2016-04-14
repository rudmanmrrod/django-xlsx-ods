from django.shortcuts import render
from django.views.generic import TemplateView,View
from django.core.files import File
from .forms import UploadFileForm
from .models import FileSave
from openpyxl import load_workbook
from ODSReader import ODSReader

# Create your views here.
class index(TemplateView):
    template_name = 'envio/index.html'
    
class envioUpload(TemplateView):
    template_name = 'envio/envio_upload.html'
    
    def get(self,request):
        form = UploadFileForm()
        return render(request,self.template_name, {'form': form})
    
    def post(self,request):
        error = ''
        form = UploadFileForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['cargar']
            extension = str(archivo).split('.')[1]
            if(extension=='xlsx'):
                my_dict = self.xlsx(archivo)
                return render(request,'envio/tablaenvio.html', {'my_dict': my_dict})
            elif(extension=='ods'):
                m = FileSave(mi_archivo=request.FILES['cargar'])
                m.save()
                my_dict = self.ods(m.mi_archivo)
                return render(request,'envio/tablaenvio.html', {'my_dict': my_dict})
            else:
                print "solo se permiten archivos xlsx y ods"
                error = "solo se permiten archivos xlsx ods"
                return render(request,self.template_name, {'form': form,'error':error})
        else:
            print form.errors
        return render(request,self.template_name, {'form': form})
    
    def xlsx(self,archivo):
        wb = load_workbook(archivo)
        ws = wb.active
        filas = len(ws.rows)
        columnas = len(ws.columns)
        my_dict = {}
        for i in range(1,columnas+1):
            my_dict[ws.cell(row=1,column=i).value] = []
        
        for i in range(1,columnas+1):
            nombre = ws.cell(row=1,column=i).value
            for j in range(2,filas+1):
                my_dict[nombre].append(ws.cell(row=j,column=i).value)
        return my_dict
    
    def ods(self,archivo):
        my_file = open(str(archivo),'rb')
        doc = ODSReader(my_file, clonespannedcolumns=True)
        table = ''
        try:
            table = doc.getSheet(u'Sheet1')
        except:
            table = doc.getSheet(u'Hoja1')
        my_dict = {}
        for i in range(len(table)):
            for j in range(len(table[i])):
                if(i==0):
                    my_dict[table[i][j]] = []
                else:
                    my_dict[table[0][j]].append(table[i][j])
        print my_dict
        return my_dict